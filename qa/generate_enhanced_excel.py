import json
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
import re

# ─── 1. SETUP ─────────────────────────────────────────────────────────────
# Colors and styles
C_DARK_HEADER = "1E293B"
C_BLUE_HEADER = "2563EB"
C_LIGHT_BLUE = "EFF6FF"
C_WHITE = "FFFFFF"
C_TEXT = "334155"
C_TITLE = "0F172A"

C_PASS = "DCFCE7"
C_PASS_FONT = "166534"
C_FAIL = "FEE2E2"
C_FAIL_FONT = "991B1B"
C_NE = "F1F5F9"
C_NE_FONT = "64748B"

def hfill(hex_color): return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
def make_font(bold=False, size=10, color="000000"): return Font(name="Calibri", bold=bold, size=size, color=color)
def make_border():
    thin = Side(border_style="thin", color="CBD5E1")
    return Border(left=thin, right=thin, top=thin, bottom=thin)
def wrap_align(): return Alignment(vertical="center", wrap_text=True)
def write_cell(ws, row, col, val, bg=C_WHITE, color=C_TEXT, bold=False, align="left"):
    c = ws.cell(row=row, column=col)
    c.value = val
    c.fill = hfill(bg)
    c.font = make_font(bold=bold, color=color)
    c.border = make_border()
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

wb = openpyxl.Workbook()

# ─── 2. MANUAL TESTS SHEET ────────────────────────────────────────────────
ws_manual = wb.active
ws_manual.title = "Manual Tests"
ws_manual.sheet_view.showGridLines = False

# Import TC_DATA from existing script
import sys
import os
sys.path.append(os.getcwd())
try:
    from generate_qa_excel import TC_DATA
except Exception as e:
    print("Error importing TC_DATA:", e)
    TC_DATA = []

M_COLS = [
    ("Test Case ID", 14), ("Module", 14), ("Test Case Title", 32), ("Test Type", 12),
    ("Preconditions", 28), ("Test Steps", 42), ("Expected Result", 36),
    ("Actual Result", 36), ("Status", 12), ("Priority", 10), ("Remarks", 24)
]

for col_idx, (col_name, col_width) in enumerate(M_COLS, start=1):
    c = ws_manual.cell(row=1, column=col_idx)
    c.value = col_name
    c.fill = hfill(C_BLUE_HEADER)
    c.font = make_font(bold=True, color=C_WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_manual.column_dimensions[get_column_letter(col_idx)].width = col_width

for r_idx, tc in enumerate(TC_DATA, start=2):
    (tc_id, module, title, tc_type, precond, steps, expected, actual, status, priority, remarks) = tc
    vals = [tc_id, module, title, tc_type, precond, steps, expected, actual, status, priority, remarks]
    ws_manual.row_dimensions[r_idx].height = 60
    
    for c_idx, val in enumerate(vals, start=1):
        bg = C_LIGHT_BLUE if r_idx % 2 == 0 else C_WHITE
        font_color = C_TEXT
        bold = False
        align = "left"
        
        if c_idx == 1: bold = True; align = "center"
        if c_idx in [2, 4, 10]: align = "center"
        
        if c_idx == 9: # Status
            if status == "Pass": bg = C_PASS; font_color = C_PASS_FONT; val = "✅ Pass"; bold = True
            elif status == "Fail": bg = C_FAIL; font_color = C_FAIL_FONT; val = "❌ Fail"; bold = True
            align = "center"
            
        write_cell(ws_manual, r_idx, c_idx, val, bg, font_color, bold, align)


# ─── 3. AUTOMATED TESTS SHEET ─────────────────────────────────────────────
ws_auto = wb.create_sheet("Automated Tests")
ws_auto.sheet_view.showGridLines = False

A_COLS = [
    ("Suite", 25), ("Test Case Title", 60), ("Status", 15), ("Duration (ms)", 15)
]

for col_idx, (col_name, col_width) in enumerate(A_COLS, start=1):
    c = ws_auto.cell(row=1, column=col_idx)
    c.value = col_name
    c.fill = hfill(C_DARK_HEADER)
    c.font = make_font(bold=True, color=C_WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_auto.column_dimensions[get_column_letter(col_idx)].width = col_width

try:
    with open('playwright_report.json', 'r', encoding='utf-8') as f:
        report = json.load(f)
    
    r_idx = 2
    for suite in report.get('suites', []):
        suite_title = suite.get('title', 'Suite')
        # Sometimes suites are nested
        for sub_suite in suite.get('suites', []):
            sub_title = sub_suite.get('title', suite_title)
            for spec in sub_suite.get('specs', []):
                test_title = spec.get('title', '')
                
                # Get result
                status = "passed"
                duration = 0
                if spec.get('tests') and spec['tests'][0].get('results'):
                    last_result = spec['tests'][0]['results'][-1]
                    status = last_result.get('status', 'failed')
                    duration = last_result.get('duration', 0)
                
                bg = C_LIGHT_BLUE if r_idx % 2 == 0 else C_WHITE
                
                # Write Suite
                write_cell(ws_auto, r_idx, 1, sub_title, bg, C_TEXT, align="center")
                # Write Title
                write_cell(ws_auto, r_idx, 2, test_title, bg, C_TEXT, bold=True, align="left")
                # Write Status
                if status == "passed":
                    write_cell(ws_auto, r_idx, 3, "✅ Pass", C_PASS, C_PASS_FONT, bold=True, align="center")
                else:
                    write_cell(ws_auto, r_idx, 3, "❌ Fail", C_FAIL, C_FAIL_FONT, bold=True, align="center")
                # Write Duration
                write_cell(ws_auto, r_idx, 4, f"{duration} ms", bg, C_TEXT, align="center")
                
                ws_auto.row_dimensions[r_idx].height = 25
                r_idx += 1
except Exception as e:
    print("Error parsing automated tests:", e)
    write_cell(ws_auto, 2, 1, f"Error: {e}")

wb.save("QA_TestCases_FleetManagement_Module4_Enhanced.xlsx")
print("Saved QA_TestCases_FleetManagement_Module4_Enhanced.xlsx")
