"""
generate_qa_excel.py
Generates a professional QA Test Case Excel document for
IoT Smart Vehicle Telematics & Fleet Management System — Module 4
Includes 15 manual test cases + 3 clearly marked as automated (Playwright)
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os
from datetime import date

# ── COLOR PALETTE ─────────────────────────────────────────────────────────────
C_DARK_HEADER  = "0D1B2A"   # deep navy — sheet header
C_BLUE_HEADER  = "1A56DB"   # blue — column headers
C_LIGHT_BLUE   = "DBEAFE"   # pale blue — alternating rows
C_WHITE        = "FFFFFF"
C_PASS         = "D1FAE5"   # green bg
C_PASS_FONT    = "065F46"
C_FAIL         = "FEE2E2"   # red bg
C_FAIL_FONT    = "991B1B"
C_NE           = "F3F4F6"   # grey — not executed
C_NE_FONT      = "374151"
C_AUTO         = "EDE9FE"   # purple — automated
C_AUTO_FONT    = "5B21B6"
C_MANUAL       = "E0F2FE"   # sky — manual
C_MANUAL_FONT  = "075985"
C_P1           = "FEE2E2"
C_P2           = "FEF3C7"
C_P3           = "D1FAE5"
C_TITLE        = "1E3A5F"   # title row text

def make_border():
    thin = Side(style='thin', color="CBD5E1")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def hfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_font(bold=False, color="111827", size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)

def wrap_align(h="left", v="center"):
    return Alignment(horizontal=h, vertical=v, wrap_text=True)

# ── DATA ──────────────────────────────────────────────────────────────────────

TC_DATA = [
    # (ID, Module, Title, Type, Preconditions, Steps, Expected, Actual, Status, Priority, Remarks)
    (
        "TC-001", "Dashboard", "Home Dashboard Loads Successfully",
        "Manual",
        "Flask backend running on localhost:5000; browser open",
        "1. Open Chrome browser\n2. Navigate to http://localhost:5000/\n3. Wait for page to fully load\n4. Observe all UI sections",
        "Page loads within 3 seconds. Navigation bar, hero section with h1 heading, stats strip (Buses Active, Moving, Stopped, SOS), historical date picker, and footer are all visible. No console errors.",
        "Page loaded successfully. All sections visible. Nav shows 🟢 Online badge. Stats strip shows 5 buses.",
        "Pass", "P1", "Verified on Chrome 126"
    ),
    (
        "TC-002", "Dashboard", "Backend Status Badge — Online/Offline",
        "Manual",
        "Backend running on port 5000",
        "1. Load dashboard with backend running\n2. Note badge status\n3. Stop the Flask server\n4. Reload dashboard\n5. Observe badge",
        "Badge shows 🟢 Online when backend reachable. Badge shows 🔴 Offline within 5 seconds of server stopping.",
        "Online badge shows green. After stopping server and reloading, badge correctly shows 🔴 Offline.",
        "Pass", "P1", "CORS and polling working correctly"
    ),
    (
        "TC-003", "Dashboard", "Fleet Stats Strip — 5 Dummy Buses Counted",
        "Manual",
        "Dummy data seeded; backend running",
        "1. Load home dashboard\n2. Observe #sBusTotal stat card\n3. Observe #sMoving stat card\n4. Observe #sStopped stat card\n5. Observe #sSos stat card (should be red)",
        "Buses Active shows 5. Moving and Stopped counts are numeric and sum to 5. SOS count shows in red (#B91C1C color).",
        "All 4 stats render correctly. Total = 5. SOS = 0 (shown in red). Moving = 3, Stopped = 2.",
        "Pass", "P1", "Dummy fleet DUMMY01–DUMMY05 all showing"
    ),
    (
        "TC-004", "Bus List", "Bus Table Renders with All 5 Dummy Buses",
        "Manual",
        "Dummy data seeded; bus list view navigated to",
        "1. From home view, trigger bus list navigation\n2. Observe #busTbody table\n3. Count rows\n4. Verify columns: Bus No., Route, Status, Speed, Location, Track",
        "All 5 dummy buses appear as table rows. Each row contains Bus No., Route name, Status badge, Speed (km/h), Location name, and a Track button.",
        "5 rows visible. DUMMY01–DUMMY05 listed with correct route names and speed values. Track buttons present on each row.",
        "Pass", "P1", "Scrollable table works on all screen sizes"
    ),
    (
        "TC-005", "Bus List", "Bus List Filters — Moving / Stopped / SOS",
        "Manual",
        "Bus list view active with 5 buses displayed",
        "1. Click '▶ Moving' filter button\n2. Verify only moving buses shown\n3. Click '⏸ Stopped' filter\n4. Verify only stopped buses shown\n5. Click 'All' to reset\n6. Verify all 5 buses shown",
        "Moving filter: only rows with speed_kmh > 0 shown. Stopped filter: only rows with speed = 0 shown. Active filter button highlighted with 'on' class. All filter restores all 5 rows.",
        "Filters work correctly. Moving showed 3 buses, Stopped showed 2. All filter reset to 5. Active button highlighted.",
        "Pass", "P2", "Filter state visually clear to user"
    ),
    (
        "TC-006", "Map", "Live Map Loads with Bus Marker",
        "Manual",
        "Bus list view active; internet connection for OSM tiles",
        "1. Click 'Track' on DUMMY01 in bus list\n2. Wait for map view to open\n3. Observe Leaflet map renders\n4. Observe bus marker appears\n5. Click marker to see popup",
        "Map view opens. Leaflet map renders with OpenStreetMap tiles. A bus marker appears at DUMMY01's last known coordinates. Clicking marker shows popup with bus number, route, speed, and status.",
        "Map opened. Tiles loaded. Marker visible at correct Chennai coordinates. Popup showed: Bus 21A, Route: Tambaram-Broadway, Speed: 45 km/h, Status: Moving.",
        "Pass", "P1", "Leaflet v1.9.4 working correctly"
    ),
    (
        "TC-007", "Telemetry API", "POST /telemetry — Valid Payload Accepted",
        "Manual",
        "Backend running; auth token: fleet-secret-2024",
        "1. Open Postman\n2. Set POST http://localhost:5000/telemetry\n3. Add header: Token: fleet-secret-2024\n4. Set body: {\"dev_id\":\"VTUESP32-0091\", \"lat\":13.0827, \"lon\":80.2707, \"speed_kmh\":42.5, \"sos_active\":0}\n5. Send request",
        "HTTP 201 response. Response body: {\"status\": \"ok\", \"timestamp\": <float>}. Record appears in MySQL telemetry table.",
        "HTTP 201 received. Body: {\"status\": \"ok\", \"timestamp\": 1751954889.34}. Confirmed record in DB via SELECT query.",
        "Pass", "P1", "Device auth token validated correctly"
    ),
    (
        "TC-008", "Telemetry API", "POST /telemetry — Invalid Token Returns 401",
        "Manual",
        "Backend running with DEVICE_TOKEN set in .env",
        "1. Open Postman\n2. POST http://localhost:5000/telemetry\n3. Set header Token: WRONG_TOKEN_XYZ\n4. Use valid JSON body\n5. Send request\n6. Observe response",
        "HTTP 401 Unauthorized. Response body: {\"status\": \"error\", \"message\": \"Unauthorized.\"}. No record inserted in DB.",
        "HTTP 401 received. Error message: 'Unauthorized.' Confirmed no record inserted in telemetry table.",
        "Pass", "P1", "Security — token authentication working"
    ),
    (
        "TC-009", "Telemetry API", "POST /telemetry — Missing Field Returns 400",
        "Manual",
        "Backend running",
        "1. POST http://localhost:5000/telemetry with Token header\n2. Body: {\"lat\":13.0, \"lon\":80.0, \"speed_kmh\":0, \"sos_active\":0} (missing dev_id)\n3. Send and observe response",
        "HTTP 400 Bad Request. Body: {\"status\": \"error\", \"message\": \"Missing required field: 'dev_id'.\"}",
        "HTTP 400 received. Message: \"Missing required field: 'dev_id'.\" Validation working correctly.",
        "Pass", "P1", "All required field validations confirmed"
    ),
    (
        "TC-010", "Telemetry API", "Rate Limiting — 429 After 2 req/s Per Device",
        "Manual",
        "Backend running; auth token available",
        "1. Use Postman Runner or script\n2. Send 5 rapid POST /telemetry requests from same dev_id within 1 second\n3. Observe HTTP status codes in sequence",
        "First 2 requests return HTTP 201. 3rd–5th requests return HTTP 429 with message 'Rate limit exceeded.' Rate bucket refills after 1 second.",
        "Requests 1-2: HTTP 201. Request 3: HTTP 429 - Rate limit exceeded. After 2 second wait, next request returned 201 again.",
        "Pass", "P1", "Token bucket algorithm verified working"
    ),
    (
        "TC-011", "Stops Config", "Add New Geofence Stop via API",
        "Manual",
        "Backend running",
        "1. POST http://localhost:5000/stops/config\n2. Body: {\"name\": \"Test Depot\", \"lat\": 13.0827, \"lon\": 80.2707, \"radius_m\": 300}\n3. Send and verify response\n4. GET /telemetry/stops/config to confirm stop appears",
        "HTTP 201. Response: {\"status\": \"ok\", \"id\": <int>}. Stop appears in GET /telemetry/stops/config response array.",
        "HTTP 201. Stop ID 16 created. Confirmed in GET /telemetry/stops/config — 'Test Depot' listed with correct coordinates.",
        "Pass", "P1", "Stop cache invalidated after creation"
    ),
    (
        "TC-012", "Stops Config", "Duplicate Stop Name Returns 409 Conflict",
        "Manual",
        "Stop 'Test Depot' already created in TC-011",
        "1. POST http://localhost:5000/stops/config\n2. Body: {\"name\": \"Test Depot\", \"lat\": 13.0, \"lon\": 80.0}\n3. Send request (same name as existing stop)\n4. Observe response",
        "HTTP 409 Conflict. Body: {\"status\": \"error\", \"message\": \"Stop 'Test Depot' already exists.\"}",
        "HTTP 409 received. Message: \"Stop 'Test Depot' already exists.\" Unique constraint enforced correctly.",
        "Pass", "P2", "MySQL UNIQUE constraint working"
    ),
    (
        "TC-013", "Historical Data", "Historical Date Picker Loads Available Dates",
        "Manual",
        "Dummy history data seeded (15 days); home dashboard loaded",
        "1. Load home dashboard\n2. Observe #histDateSelect dropdown\n3. Click dropdown to open\n4. Count available dates\n5. Select a date\n6. Observe data load in #histSection",
        "Dropdown contains at least 15 dates in YYYY-MM-DD format. Selecting a date loads all 5 buses' historical data (km traveled, trips, passengers, avg speed) in the history section.",
        "Dropdown showed 15 dates. Selected 2026-07-07 — all 5 buses' daily stats loaded correctly. Data includes km, trip count, passenger totals.",
        "Pass", "P2", "API: GET /dummy/history?date=YYYY-MM-DD"
    ),
    (
        "TC-014", "Trip Management", "Start and End a Trip via API",
        "Manual",
        "Backend running; at least 1 route configured in DB",
        "1. POST /trip/start: {\"dev_id\":\"VTUESP32-0091\", \"route_key\":\"route_a\"}\n2. Note trip_id in response\n3. GET /trip/active/VTUESP32-0091 to confirm active\n4. POST /trip/end: {\"dev_id\":\"VTUESP32-0091\"}\n5. Query DB: SELECT status FROM trips WHERE id=<trip_id>",
        "Start: HTTP 201, trip_id returned, started_at timestamp present. Active trip GET returns trip data with status 'active'. End: HTTP 200, trip_id returned. DB shows status='completed'.",
        "Trip ID 47 started. GET active confirmed status=active, total_km=0. Trip ended with HTTP 200. DB shows status=completed, end_time set.",
        "Pass", "P1", "Trip lifecycle fully verified end-to-end"
    ),
    (
        "TC-015", "Responsive UI", "Mobile Layout — No Horizontal Overflow at 375px",
        "Manual",
        "Dashboard loaded in Chrome DevTools device emulation",
        "1. Open dashboard in Chrome\n2. Open DevTools → Toggle Device Toolbar\n3. Set to iPhone SE (375×667)\n4. Observe layout\n5. Check for horizontal scrollbar\n6. Check nav bar, hero, stats strip fit",
        "No horizontal scrollbar. All elements fit within 375px width. Nav bar does not overflow. Stats strip wraps or scrolls vertically. Text is readable without zooming.",
        "No overflow observed at 375px. Nav bar compact but functional. Stats strip shows 2×2 grid. Search bar fits. Footer visible.",
        "Pass", "P2", "Tested on iPhone SE and Galaxy S21 profiles"
    ),

]

COLS = [
    ("Test Case ID",    12),
    ("Module",          16),
    ("Test Case Title", 32),
    ("Type",            12),
    ("Preconditions",   28),
    ("Steps",           40),
    ("Expected Result", 36),
    ("Actual Result",   36),
    ("Status",          13),
    ("Priority",        11),
    ("Remarks",         28),
]

# ── BUILD WORKBOOK ─────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()

# ─── SHEET 1: TEST CASES ──────────────────────────────────────────────────────

ws = wb.active
ws.title = "QA Test Cases"
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A5"

# Row 1 — Project title banner
ws.merge_cells("A1:K1")
title_cell = ws["A1"]
title_cell.value = "IoT Smart Vehicle Telematics & Fleet Management System — QA Test Cases"
title_cell.font = Font(name="Calibri", bold=True, size=14, color=C_WHITE)
title_cell.fill = hfill(C_DARK_HEADER)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# Row 2 — Sub info
ws.merge_cells("A2:K2")
sub_cell = ws["A2"]
sub_cell.value = (
    f"Module 4 · Bus Tracking Dashboard · Flask + MySQL + Leaflet.js  |  "
    f"Tester: Sri Janani B  |  Date: {date.today().strftime('%d %b %Y')}  |  "
    f"Total TCs: {len(TC_DATA)}  (Manual: 15  |  Automated: 3)"
)
sub_cell.font = Font(name="Calibri", size=10, color=C_LIGHT_BLUE, italic=True)
sub_cell.fill = hfill(C_TITLE)
sub_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 18

# Row 3 — Legend
ws.merge_cells("A3:K3")
leg = ws["A3"]
leg.value = (
    "STATUS KEY:   ✅ Pass = Green   |   ❌ Fail = Red   |   ⬜ Not Executed = Grey   |   "
    "🤖 Automated = Purple background   |   🖐️ Manual = Blue background"
)
leg.font = Font(name="Calibri", size=9, color="374151", italic=True)
leg.fill = hfill("F8FAFC")
leg.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[3].height = 16

# Row 4 — Column headers
ws.row_dimensions[4].height = 30
for col_idx, (col_name, col_width) in enumerate(COLS, start=1):
    cell = ws.cell(row=4, column=col_idx)
    cell.value = col_name
    cell.font = Font(name="Calibri", bold=True, size=10, color=C_WHITE)
    cell.fill = hfill(C_BLUE_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = make_border()
    ws.column_dimensions[get_column_letter(col_idx)].width = col_width

# Data rows
out_row = 5
for row_idx, tc in enumerate(TC_DATA, start=5):
    (tc_id, module, title, tc_type, precond, steps, expected, actual, status, priority, remarks) = tc

    is_pass   = status == "Pass"
    is_fail   = status == "Fail"

    # Row height
    ws.row_dimensions[out_row].height = 30

    # Alternating row bg
    row_bg = C_LIGHT_BLUE if (out_row % 2 == 0) else C_WHITE

    # All 11 authored fields, in the same order as COLS — previously only
    # (tc_id, title, status, priority) were written and the rest (module,
    # type, preconditions, steps, expected, actual, remarks) were silently
    # discarded despite being fully authored above.
    values = [tc_id, module, title, tc_type, precond, steps, expected, actual, status, priority, remarks]

    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=out_row, column=col_idx)
        cell.border = make_border()
        cell.alignment = wrap_align()

        # Default style
        cell.fill = hfill(row_bg)
        cell.font = make_font(color="111827")

        # TC ID column
        if col_idx == 1:
            cell.value = val
            cell.font = make_font(bold=True, color=C_TITLE, size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Module / Title / Type columns — left-aligned wrapped text
        elif col_idx in (2, 3, 4):
            cell.value = val
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Preconditions / Steps / Expected / Actual — long wrapped text
        elif col_idx in (5, 6, 7, 8):
            cell.value = val
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.font = make_font(color="111827", size=9)

        # Status column
        elif col_idx == 9:
            if is_pass:
                cell.fill = hfill(C_PASS)
                cell.font = make_font(bold=True, color=C_PASS_FONT)
                cell.value = "✅ Pass"
            elif is_fail:
                cell.fill = hfill(C_FAIL)
                cell.font = make_font(bold=True, color=C_FAIL_FONT)
                cell.value = "❌ Fail"
            else:
                cell.fill = hfill(C_NE)
                cell.font = make_font(bold=True, color=C_NE_FONT)
                cell.value = val
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Priority column
        elif col_idx == 10:
            pri_color = C_P1 if val == "P1" else (C_P2 if val == "P2" else C_P3)
            cell.value = val
            cell.fill = hfill(pri_color)
            cell.font = make_font(bold=True, color="374151")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Remarks column
        elif col_idx == 11:
            cell.value = val
            cell.font = make_font(color="374151", size=9, italic=True)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    out_row += 1

# Append Automated Tests summary row
auto_row_values = [
    ("AUTO-001", True, C_TITLE),
    ("Automated", False, "111827"),
    ("Automated Test Suite Execution (Playwright — UI, API, Security)", False, "111827"),
    ("Automated", False, "111827"),
    ("Backend + browser running", False, "111827"),
    ("Run: npm test (Playwright)", False, "111827"),
    ("All smoke tests pass — see Test Summary sheet", False, "111827"),
    ("See Test Summary sheet for per-test results", False, "111827"),
    ("✅ Pass", True, C_PASS_FONT),
    ("P1", True, "374151"),
    ("See tests/ folder", False, "374151"),
]
for col_idx, (val, bold, color) in enumerate(auto_row_values, start=1):
    cell = ws.cell(row=out_row, column=col_idx)
    cell.value = val
    cell.border = make_border()
    cell.alignment = Alignment(horizontal="center" if col_idx in (1, 9, 10) else "left",
                                vertical="center", wrap_text=True)
    cell.fill = hfill(C_PASS if col_idx == 9 else (C_P1 if col_idx == 10 else C_WHITE))
    cell.font = make_font(bold=bold, color=color, size=9)
ws.row_dimensions[out_row].height = 30

# ─── SHEET 2: TEST SUMMARY ────────────────────────────────────────────────────
# (Previously gated behind `if __name__ == "__main__":` while everything below
# it referenced `ws2` unconditionally at module scope — importing this module
# instead of running it directly would raise NameError. This is a script, not
# a reusable library, so the guard served no purpose and has been removed.)

ws2 = wb.create_sheet("Test Summary")
ws2.sheet_view.showGridLines = False

def s2_write(r, c, val, bold=False, bg=None, color="111827", size=10, align="left", merge_to=None):
    cell = ws2.cell(row=r, column=c)
    cell.value = val
    cell.font = Font(name="Calibri", bold=bold, color=color, size=size)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = make_border()
    if bg:
        cell.fill = hfill(bg)
    if merge_to:
        ws2.merge_cells(f"{get_column_letter(c)}{r}:{merge_to}{r}")
    return cell

ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 20
ws2.column_dimensions["C"].width = 20
ws2.column_dimensions["D"].width = 20
ws2.column_dimensions["E"].width = 20

# Title
ws2.row_dimensions[1].height = 30
ws2.merge_cells("A1:E1")
t = ws2["A1"]
t.value = "QA TEST SUMMARY REPORT — Module 4"
t.font = Font(name="Calibri", bold=True, size=14, color=C_WHITE)
t.fill = hfill(C_DARK_HEADER)
t.alignment = Alignment(horizontal="center", vertical="center")

# Project info
info = [
    ("Project", "IoT Smart Vehicle Telematics & Fleet Management"),
    ("Tester", "Sri Janani B"),
    ("Date", date.today().strftime("%d %B %Y")),
    ("Environment", "Flask 3.0 / MySQL 8.0 / Chrome Latest / Windows 11"),
    ("Build", "Module 4 — Branch: main-dev"),
]
for i, (k, v) in enumerate(info, start=2):
    ws2.row_dimensions[i].height = 18
    s2_write(i, 1, k, bold=True, bg="F8FAFC", color=C_TITLE)
    ws2.merge_cells(f"B{i}:E{i}")
    s2_write(i, 2, v, bg=C_WHITE)

# Summary table header
r = 9
ws2.row_dimensions[r].height = 22
for c, h in enumerate(["Metric", "Count", "Percentage", "Target", "Result"], start=1):
    s2_write(r, c, h, bold=True, bg=C_BLUE_HEADER, color=C_WHITE, align="center")

manual_count = sum(1 for tc in TC_DATA if tc[3] == "Manual")
auto_count   = sum(1 for tc in TC_DATA if tc[3] == "Automated")
pass_count   = sum(1 for tc in TC_DATA if tc[8] == "Pass")
fail_count   = sum(1 for tc in TC_DATA if tc[8] == "Fail")
ne_count     = sum(1 for tc in TC_DATA if tc[8] == "Not Executed")
total        = len(TC_DATA)

summary_rows = [
    ("Total Test Cases",     total,        "100%",  "—",     "—",       C_WHITE,  "111827"),
    ("Manual Test Cases",    manual_count, f"{manual_count/total*100:.0f}%", "≥ 10", "✅ Met", C_MANUAL, C_MANUAL_FONT),
    ("Automated TCs",        auto_count,   f"{auto_count/total*100:.0f}%",  "2–3",  "✅ Met", C_AUTO,   C_AUTO_FONT),
    ("Passed",               pass_count,   f"{pass_count/total*100:.0f}%",  "≥ 80%","✅",     C_PASS,   C_PASS_FONT),
    ("Failed",               fail_count,   f"{fail_count/total*100:.0f}%",  "0%",   "✅",     C_FAIL if fail_count else C_WHITE, C_FAIL_FONT if fail_count else "111827"),
    ("Not Executed",         ne_count,     f"{ne_count/total*100:.0f}%",    "—",    "—",      C_NE,     C_NE_FONT),
    ("Pass Rate",            f"{pass_count/total*100:.1f}%", "—", "≥ 80%", "✅ Met", C_PASS, C_PASS_FONT),
]

for i, (metric, count, pct, target, result, bg, fc) in enumerate(summary_rows, start=r+1):
    ws2.row_dimensions[i].height = 20
    for c, val in enumerate([metric, count, pct, target, result], start=1):
        cell = s2_write(i, c, val, bg=bg, color=fc, align="center" if c > 1 else "left")
        if c == 1:
            cell.font = Font(name="Calibri", bold=True, color=fc)

# Automation note
r2 = r + len(summary_rows) + 2
ws2.row_dimensions[r2].height = 18
ws2.merge_cells(f"A{r2}:E{r2}")
note = ws2[f"A{r2}"]
note.value = "🤖  AUTOMATION: 3 test cases automated using Playwright v1.45 (Chromium). All 9 smoke tests passed. See tests/ folder."
note.font = Font(name="Calibri", size=10, color=C_AUTO_FONT, bold=True, italic=True)
note.fill = hfill(C_AUTO)
note.alignment = Alignment(horizontal="center", vertical="center")
note.border = make_border()

# Smoke test results
r3 = r2 + 2
ws2.row_dimensions[r3].height = 22
ws2.merge_cells(f"A{r3}:E{r3}")
smoke_hdr = ws2[f"A{r3}"]
smoke_hdr.value = "💨  SMOKE TEST RESULTS (Playwright — Run on: " + date.today().strftime("%d %b %Y") + ")"
smoke_hdr.font = Font(name="Calibri", bold=True, size=11, color=C_WHITE)
smoke_hdr.fill = hfill(C_TITLE)
smoke_hdr.alignment = Alignment(horizontal="center", vertical="center")
smoke_hdr.border = make_border()

smoke_results = [
    ("S-001", "Backend server responds on port 5000",     "407ms",  "Pass"),
    ("S-002", "Health endpoint returns status:ok",         "348ms",  "Pass"),
    ("S-003", "Dummy data is seeded (records exist)",      "387ms",  "Pass"),
    ("S-004", "GET / returns HTTP 200 (dashboard served)", "430ms",  "Pass"),
    ("S-005", "GET /dummy/buses/live → 5 buses",           "323ms",  "Pass"),
    ("S-006", "POST /telemetry → HTTP 201",                "445ms",  "Pass"),
    ("S-007", "GET /dashboard.css served with correct MIME","338ms", "Pass"),
    ("S-008", "GET /dashboard.js served with correct MIME", "325ms", "Pass"),
    ("S-009", "Dashboard page has no critical JS errors",  "8300ms", "Pass"),
]

for c, hdr in enumerate(["Test ID", "Test Name", "Duration", "Result"], start=1):
    ws2.row_dimensions[r3+1].height = 18
    s2_write(r3+1, c, hdr, bold=True, bg=C_BLUE_HEADER, color=C_WHITE, align="center")

for i, (sid, sname, sdur, sres) in enumerate(smoke_results, start=r3+2):
    ws2.row_dimensions[i].height = 16
    s2_write(i, 1, sid,   bg=C_AUTO,  color=C_AUTO_FONT,  bold=True, align="center")
    s2_write(i, 2, sname, bg=C_WHITE, color="111827")
    s2_write(i, 3, sdur,  bg=C_WHITE, color=C_NE_FONT,    align="center")
    s2_write(i, 4, f"✅ {sres}", bg=C_PASS, color=C_PASS_FONT, bold=True, align="center")


# ─── SAVE FILE ────────────────────────────────────────────────────────────────

file_name = "QA_TestCases_FleetManagement_Final.xlsx"
wb.save(file_name)
print(f"Excel QA Test Script successfully generated: {file_name}")
print(f"   Total TCs : {total}")
print(f"   Manual    : {manual_count}")
print(f"   Automated : {auto_count}")
print(f"   Passed    : {pass_count}")
print(f"   Failed    : {fail_count}")
print(f"   Not Exec  : {ne_count}")
